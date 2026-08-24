import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..database import get_db
from .. import models, schemas
from ..crypto import encrypt_payload

router = APIRouter(prefix="/api/assets", tags=["assets"])


@router.post("/", response_model=schemas.AssetResponse)
def create_asset(asset: schemas.AssetCreate, db: Session = Depends(get_db)):
    db_asset = db.query(models.Asset).filter(models.Asset.asset_code == asset.asset_code).first()
    if db_asset:
        raise HTTPException(status_code=400, detail="Asset code already exists")
    
    db_asset = models.Asset(**asset.dict())
    db.add(db_asset)
    db.commit()
    db.refresh(db_asset)
    return db_asset


@router.get("/", response_model=List[schemas.AssetResponse])
def list_assets(
    skip: int = 0,
    limit: int = 100,
    department: Optional[str] = None,
    status: Optional[str] = None,
    location: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.Asset)
    if department:
        query = query.filter(models.Asset.department == department)
    if status:
        query = query.filter(models.Asset.status == status)
    if location:
        query = query.filter(models.Asset.location.contains(location))
    return query.offset(skip).limit(limit).all()


@router.get("/rfid/{rfid_tag}", response_model=schemas.AssetResponse)
def get_asset_by_rfid(rfid_tag: str, db: Session = Depends(get_db)):
    asset = db.query(models.Asset).filter(models.Asset.rfid_tag == rfid_tag).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return asset


@router.get("/{asset_code}", response_model=schemas.AssetResponse)
def get_asset(asset_code: str, db: Session = Depends(get_db)):
    asset = db.query(models.Asset).filter(models.Asset.asset_code == asset_code).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return asset


@router.put("/{asset_code}", response_model=schemas.AssetResponse)
def update_asset(asset_code: str, asset_update: schemas.AssetUpdate, db: Session = Depends(get_db)):
    asset = db.query(models.Asset).filter(models.Asset.asset_code == asset_code).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    update_data = asset_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(asset, field, value)
    
    db.commit()
    db.refresh(asset)
    return asset


@router.delete("/{asset_code}")
def delete_asset(asset_code: str, db: Session = Depends(get_db)):
    asset = db.query(models.Asset).filter(models.Asset.asset_code == asset_code).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    db.delete(asset)
    db.commit()
    return {"message": "Asset deleted"}


@router.post("/{asset_code}/qr-token", response_model=schemas.QRTokenResponse)
def generate_qr_token(asset_code: str, db: Session = Depends(get_db)):
    asset = db.query(models.Asset).filter(models.Asset.asset_code == asset_code).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    token = encrypt_payload(asset_code)
    # URL will be resolved by frontend based on current host
    return {
        "token": token,
        "url": f"/s/{token}",
        "asset_code": asset_code
    }


@router.post("/batch-qr-tokens")
def batch_qr_tokens(asset_codes: List[str], db: Session = Depends(get_db)):
    tokens = []
    for code in asset_codes:
        asset = db.query(models.Asset).filter(models.Asset.asset_code == code).first()
        if asset:
            token = encrypt_payload(code)
            tokens.append({
                "token": token,
                "url": f"/s/{token}",
                "asset_code": code,
                "asset_name": asset.asset_name,
                "location": asset.location
            })
    return {"tokens": tokens}


# 中文列名到模型字段的映射
COLUMN_MAP = {
    "资产编号": "asset_code", "资产名称": "asset_name", "管理单元": "management_unit",
    "品牌": "brand", "规格型号": "spec_model", "出厂编号": "serial_no",
    "取得方式": "acquisition_method", "购建日期": "purchase_date", "启用时间": "book_date",
    "使用单位": "department", "使用部门": "department", "管理者": "responsible_person",
    "使用者": "current_user", "使用地点": "location", "归口管理": "custodian_dept",
    "采购部门": "procurement_dept", "资产原值": "purchase_price", "牌照号": "license_plate",
    "出厂日期": "factory_date", "制造厂家": "manufacturer", "供应商": "supplier",
    "备注": "remarks", "设备代码": "equipment_code", "设备成组": "equipment_group",
    "决算出项": "final_account", "建卡时间": "card_date", "账套": "account_book",
    "凭证号": "voucher_no", "核算项目": "accounting_item", "核算科目": "accounting_subject",
    "主体资产": "purchase_price", "增值税": "vat", "安装费": "installation_fee",
    "运杂费": "shipping_fee", "其他费用": "other_fee", "折旧状态": "depreciation_status",
    "折旧方法": "depreciation_method", "折旧期限(月)": "useful_life_months",
    "残值率(%)": "residual_rate", "月折旧率": "monthly_depreciation_rate",
    "月折旧额": "monthly_depreciation_amount", "合同代码": "contract_code",
    "验收单": "acceptance_doc", "调拨单": "transfer_doc", "可扫描": "scanable",
    "资产类码": "asset_class_code", "分类码描述": "category", "状态": "status",
    "RFID标签号": "rfid_tag",
    # 兼容原有英文列名
    "asset_code": "asset_code", "asset_name": "asset_name", "category": "category",
    "spec_model": "spec_model", "serial_no": "serial_no", "brand": "brand",
    "purchase_date": "purchase_date", "purchase_price": "purchase_price",
    "department": "department", "current_user": "current_user", "location": "location",
    "status": "status", "supplier": "supplier", "useful_life_months": "useful_life_months",
    "depreciation_method": "depreciation_method", "accumulated_depreciation": "accumulated_depreciation",
    "net_value": "net_value", "responsible_person": "responsible_person",
    "book_date": "book_date", "voucher_no": "voucher_no", "remarks": "remarks",
    "rfid_tag": "rfid_tag", "management_unit": "management_unit",
    "acquisition_method": "acquisition_method", "custodian_dept": "custodian_dept",
    "procurement_dept": "procurement_dept", "license_plate": "license_plate",
    "factory_date": "factory_date", "manufacturer": "manufacturer",
    "equipment_code": "equipment_code", "equipment_group": "equipment_group",
    "final_account": "final_account", "card_date": "card_date",
    "account_book": "account_book", "accounting_item": "accounting_item",
    "accounting_subject": "accounting_subject", "vat": "vat",
    "installation_fee": "installation_fee", "shipping_fee": "shipping_fee",
    "other_fee": "other_fee", "depreciation_status": "depreciation_status",
    "residual_rate": "residual_rate", "monthly_depreciation_rate": "monthly_depreciation_rate",
    "monthly_depreciation_amount": "monthly_depreciation_amount",
    "contract_code": "contract_code", "acceptance_doc": "acceptance_doc",
    "transfer_doc": "transfer_doc", "scanable": "scanable",
    "asset_class_code": "asset_class_code",
}

DATE_FIELDS = {"purchase_date", "book_date", "factory_date", "card_date"}
NUMERIC_FIELDS = {"purchase_price", "accumulated_depreciation", "net_value", "vat",
                  "installation_fee", "shipping_fee", "other_fee", "residual_rate",
                  "monthly_depreciation_rate", "monthly_depreciation_amount"}
INT_FIELDS = {"useful_life_months"}


@router.post("/import", response_model=schemas.ImportResponse)
def import_assets(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")

    contents = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    required_fields = {"asset_code", "asset_name", "category", "purchase_date", "purchase_price", "department", "location", "responsible_person"}

    errors = []
    success = 0
    failed = 0

    # Build column mapping from header row
    col_map = {}
    for idx, h in enumerate(headers):
        if h and h in COLUMN_MAP:
            col_map[COLUMN_MAP[h]] = idx

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = {}
            for field, col_idx in col_map.items():
                val = row[col_idx]
                if field in DATE_FIELDS and val:
                    if isinstance(val, str):
                        from datetime import datetime
                        val = datetime.strptime(val.strip(), "%Y-%m-%d").date()
                if field in NUMERIC_FIELDS and val:
                    val = float(val)
                if field in INT_FIELDS and val:
                    val = int(val)
                # 状态映射: 已启用 -> 在用
                if field == "status" and val:
                    val = str(val).strip()
                    if val == "已启用":
                        val = "在用"
                data[field] = val

            # Check required
            for rf in required_fields:
                if rf not in data or not data[rf]:
                    raise ValueError(f"Missing required field: {rf}")

            # Check exists
            existing = db.query(models.Asset).filter(models.Asset.asset_code == data["asset_code"]).first()
            if existing:
                for k, v in data.items():
                    setattr(existing, k, v)
            else:
                asset = models.Asset(**data)
                db.add(asset)

            success += 1
        except Exception as e:
            failed += 1
            errors.append(f"Row {row_idx}: {str(e)}")

    db.commit()
    return schemas.ImportResponse(success=success, failed=failed, errors=errors)


@router.get("/template/download")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "固定资产卡片导入模板"

    headers = [
        "资产编号", "资产名称", "管理单元", "品牌", "规格型号", "出厂编号",
        "取得方式", "购建日期", "启用时间", "使用单位", "使用部门", "管理者",
        "使用者", "使用地点", "归口管理", "采购部门", "资产原值", "牌照号",
        "出厂日期", "制造厂家", "供应商", "备注", "设备代码", "设备成组",
        "决算出项", "建卡时间", "账套", "凭证号", "核算项目", "核算科目",
        "主体资产", "增值税", "安装费", "运杂费", "其他费用", "折旧状态",
        "折旧方法", "折旧期限(月)", "残值率(%)", "月折旧率", "月折旧额",
        "合同代码", "验收单", "调拨单", "可扫描", "资产类码", "分类码描述",
        "状态", "RFID标签号"
    ]

    ws.append(headers)

    sample = [
        "1011048004132", "华为信创台式机电脑", "三峡物资招标管理有限公司(800)", "华为", "擎云W515", "0412230727006759",
        "新购", "2026-08-06", "2026-08-06", "三峡物资招标公司(510)", "三峡物资招标公司抽水蓄能业务部(51022)", "李刚",
        "李刚", "湖北省襄阳市南漳县李庙镇闫坪村南漳抽水蓄能项目部", "", "", 3880.84, "",
        "", "", "", "远达物流资产转让设备", "", "",
        "", "2026-08-06", "三峡物资招标公司帐套", "", "", "固定资产",
        3880.84, 0, "", "", "", "折旧",
        "线性折旧", "60", "0", "", "",
        "", "", "", "否", "101104", "台式计算机",
        "已启用", ""
    ]
    ws.append(sample)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=asset_import_template.xlsx"}
    )
