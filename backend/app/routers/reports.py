import io
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ..database import get_db
from .. import models, schemas

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/diff/{task_id}", response_model=schemas.DiffReportResponse)
def generate_diff_report(task_id: str, status_filter: Optional[str] = None, db: Session = Depends(get_db)):
    task = db.query(models.InventoryTask).filter(models.InventoryTask.task_id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    all_assets = db.query(models.Asset).all()
    checked_records = db.query(models.InventoryRecord).filter(
        models.InventoryRecord.task_id == task_id
    ).all()

    checked_map = {r.asset_code: r for r in checked_records}

    diff_items = []
    for asset in all_assets:
        record = checked_map.get(asset.asset_code)

        if record:
            diff_type = "已盘点"
        else:
            diff_type = "未盘点"

        diff_items.append(schemas.DiffReportItem(
            asset_code=asset.asset_code,
            asset_name=asset.asset_name,
            diff_type=diff_type,
            book_value=asset.net_value or asset.purchase_price,
            expected_location=asset.location,
            actual_location=record.actual_location if record else None,
            expected_status=asset.status.value if asset.status else None,
            actual_status=None,
            inventory_time=record.inventory_time if record else None,
            inventory_person=record.inventory_person if record else None
        ))

    # Check for extra assets checked but not in system
    asset_codes = {a.asset_code for a in all_assets}
    for record in checked_records:
        if record.asset_code not in asset_codes:
            diff_items.append(schemas.DiffReportItem(
                asset_code=record.asset_code,
                asset_name="未知资产",
                diff_type="已盘点",
                book_value=None,
                expected_location=None,
                actual_location=record.actual_location,
                expected_status=None,
                actual_status=None,
                inventory_time=record.inventory_time,
                inventory_person=record.inventory_person
            ))

    if status_filter and status_filter != "全部":
        diff_items = [i for i in diff_items if i.diff_type == status_filter]

    return schemas.DiffReportResponse(
        task_id=task_id,
        task_name=task.task_name,
        total_assets=len(all_assets),
        checked_assets=len(checked_records),
        diff_items=diff_items,
        generated_at=datetime.now()
    )


@router.get("/diff/{task_id}/export")
def export_diff_report(task_id: str, status_filter: Optional[str] = None, db: Session = Depends(get_db)):
    report = generate_diff_report(task_id, status_filter, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "盘点报告"

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = f"固定资产盘点报告 - {report.task_name or task_id}"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Summary
    ws["A3"] = f"任务ID: {task_id}"
    ws["A4"] = f"资产总数: {report.total_assets}"
    ws["A5"] = f"已盘点数: {report.checked_assets}"
    ws["A6"] = f"生成时间: {report.generated_at.strftime('%Y-%m-%d %H:%M:%S')}"

    # Headers
    headers = ["资产编号", "资产名称", "盘点情况", "账面净值", "台账地点", "实际地点", "台账状态", "实际状态", "实际盘点时间", "盘点人员"]
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    type_colors = {
        "已盘点": "C6EFCE",
        "未盘点": "FFC7CE"
    }

    for item in report.diff_items:
        row = [
            item.asset_code,
            item.asset_name,
            item.diff_type,
            float(item.book_value) if item.book_value else "",
            item.expected_location or "",
            item.actual_location or "",
            item.expected_status or "",
            item.actual_status or "",
            item.inventory_time.strftime('%Y-%m-%d %H:%M:%S') if item.inventory_time else "",
            item.inventory_person or ""
        ]
        ws.append(row)

        # Color coding
        fill_color = type_colors.get(item.diff_type, "FFFFFF")
        for col in range(1, 11):
            ws.cell(row=ws.max_row, column=col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=inventory_report_{task_id}.xlsx"}
    )
